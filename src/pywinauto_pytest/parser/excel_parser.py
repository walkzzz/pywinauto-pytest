from openpyxl import load_workbook
from typing import List, Dict, Any
from .base_parser import BaseParser
from ..models import TestSuite, TestCase, TestStep


class ExcelParser(BaseParser):
    """Excel 格式测试用例解析器"""
    
    def parse(self, file_path: str) -> TestSuite:
        """解析 Excel 文件，返回 TestSuite 对象"""
        workbook = load_workbook(file_path)
        sheet = workbook.active
        
        # 解析测试套件信息（假设在 A1 单元格）
        test_suite_name = sheet['A1'].value if sheet['A1'].value else 'Unnamed Suite'
        test_suite = TestSuite(name=test_suite_name, tests=[])
        
        # 解析测试用例
        test_cases = self._parse_test_cases(sheet)
        test_suite.tests = test_cases
        
        return test_suite
    
    def supports(self, file_path: str) -> bool:
        """检查是否支持 Excel 文件"""
        ext = self._get_file_extension(file_path)
        return ext in ['xlsx', 'xls']
    
    def _parse_test_cases(self, sheet) -> List[TestCase]:
        """解析测试用例"""
        test_cases = []
        current_test = None
        current_steps = []
        current_teardown = []
        
        # 假设测试用例从第 3 行开始
        for row in sheet.iter_rows(min_row=3, values_only=True):
            # 跳过空行
            if not any(row):
                continue
            
            # 解析测试用例名称（以 "## " 开头）
            if row[0] and isinstance(row[0], str) and row[0].startswith('## '):
                # 如果已有未完成的测试用例，保存它
                if current_test:
                    current_test.steps = current_steps
                    current_test.teardown = current_teardown
                    test_cases.append(current_test)
                    current_steps = []
                    current_teardown = []
                
                # 创建新的测试用例
                test_name = row[0][3:]
                current_test = TestCase(name=test_name, steps=[])
            
            # 解析测试步骤
            elif current_test and row[1] and isinstance(row[1], str):
                action = row[1]
                target = row[2] if row[2] else None
                locator = row[3] if row[3] else None
                expected = row[4] if row[4] else None
                
                step = TestStep(
                    action=action,
                    target=target,
                    locator=locator,
                    expected=expected
                )
                
                # 判断是主要步骤还是 teardown 步骤
                if row[0] and isinstance(row[0], str) and row[0].startswith('Teardown'):
                    current_teardown.append(step)
                else:
                    current_steps.append(step)
        
        # 保存最后一个测试用例
        if current_test:
            current_test.steps = current_steps
            current_test.teardown = current_teardown
            test_cases.append(current_test)
        
        return test_cases
